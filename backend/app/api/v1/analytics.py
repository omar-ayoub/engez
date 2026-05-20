import csv
import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import get_tenant_scope, require_accountant_or_admin
from app.core.tenant import TenantScope
from app.models.user import User
from app.schemas.analytics import SpendByProject, SpendByCategory, SpendTrend, BudgetVsActual, AnalyticsSummary
from app.services.analytics import spend_by_project, spend_by_category, spend_trend, budget_vs_actual, summary

router = APIRouter(prefix="/analytics", tags=["analytics"])


@router.get("/spend-by-project", response_model=list[SpendByProject])
async def get_spend_by_project(
    days: int = Query(30, ge=7, le=365),
    scope: TenantScope = Depends(get_tenant_scope),
    user: User = Depends(require_accountant_or_admin),
    db: AsyncSession = Depends(get_db),
):
    return await spend_by_project(db, scope.company_id, days)


@router.get("/spend-by-category", response_model=list[SpendByCategory])
async def get_spend_by_category(
    days: int = Query(30, ge=1, le=365),
    scope: TenantScope = Depends(get_tenant_scope),
    user: User = Depends(require_accountant_or_admin),
    db: AsyncSession = Depends(get_db),
):
    return await spend_by_category(db, scope.company_id, days)


@router.get("/spend-trend", response_model=list[SpendTrend])
async def get_spend_trend(
    days: int = Query(90, ge=7, le=365),
    scope: TenantScope = Depends(get_tenant_scope),
    user: User = Depends(require_accountant_or_admin),
    db: AsyncSession = Depends(get_db),
):
    return await spend_trend(db, scope.company_id, days)


@router.get("/budget-vs-actual", response_model=list[BudgetVsActual])
async def get_budget_vs_actual(
    scope: TenantScope = Depends(get_tenant_scope),
    user: User = Depends(require_accountant_or_admin),
    db: AsyncSession = Depends(get_db),
):
    return await budget_vs_actual(db, scope.company_id)


@router.get("/summary", response_model=AnalyticsSummary)
async def get_summary(
    days: int = Query(30, ge=1, le=365),
    scope: TenantScope = Depends(get_tenant_scope),
    user: User = Depends(require_accountant_or_admin),
    db: AsyncSession = Depends(get_db),
):
    return await summary(db, scope.company_id, days)


@router.get("/export")
async def export_analytics(
    format: str = Query(...),
    view: str = Query(...),
    days: int = Query(30, ge=1, le=365),
    scope: TenantScope = Depends(get_tenant_scope),
    user: User = Depends(require_accountant_or_admin),
    db: AsyncSession = Depends(get_db),
):
    if format not in ("csv", "excel"):
        raise HTTPException(status_code=422, detail="Invalid format. Use 'csv' or 'excel'.")
    if view not in ("spend-by-project", "spend-by-category", "spend-trend", "budget-vs-actual"):
        raise HTTPException(status_code=422, detail="Invalid view.")

    data = await _get_view_data(db, scope.company_id, view, days)

    if format == "csv":
        return _csv_response(data, view, days)
    return _excel_response(data, view, days)


async def _get_view_data(db, company_id: str, view: str, days: int) -> list[dict]:
    if view == "spend-by-project":
        return await spend_by_project(db, company_id, days)
    if view == "spend-by-category":
        return await spend_by_category(db, company_id, days)
    if view == "spend-trend":
        return await spend_trend(db, company_id, days)
    if view == "budget-vs-actual":
        return await budget_vs_actual(db, company_id)
    return []


def _csv_response(data: list[dict], view: str, days: int) -> StreamingResponse:
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=analytics-{view}-{days}d.csv"},
    )


def _excel_response(data: list[dict], view: str, days: int) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = view

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    if data:
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, item in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col, value=item.get(header, ""))
    else:
        ws.cell(row=1, column=1, value="No data")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=analytics-{view}-{days}d.xlsx"},
    )
